import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

# Initialize session state for tracking processing status
if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False


def copy_cell_formatting(source_cell, target_cell):
    """Copy formatting from source cell to target cell"""
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color,
        )
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color,
        )
        target_cell.border = Border(
            left=Side(border_style=source_cell.border.left.style),
            right=Side(border_style=source_cell.border.right.style),
            top=Side(border_style=source_cell.border.top.style),
            bottom=Side(border_style=source_cell.border.bottom.style),
        )


def adjust_formula_row(formula, row_diff):
    """Adjust formula references for new row"""
    new_formula = formula
    import re

    cell_refs = re.findall(r"([A-Z]+)(\d+)", formula)
    for col, row in cell_refs:
        old_ref = f"{col}{row}"
        new_row = int(row) + row_diff
        new_ref = f"{col}{new_row}"
        new_formula = new_formula.replace(old_ref, new_ref)
    return new_formula


def copy_formulas(template_ws, source_row, target_row):
    """Copy formulas from source row to target row, adjusting cell references"""
    row_diff = target_row - source_row
    for cell in template_ws[source_row]:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            target_cell = template_ws.cell(row=target_row, column=cell.column)
            adjusted_formula = adjust_formula_row(cell.value, row_diff)
            target_cell.value = adjusted_formula


def process_profile_data(uploaded_profile_df, template_file):
    """Process profile data and update template"""
    # Handle both file object and file path
    if isinstance(template_file, str):
        template_wb = openpyxl.load_workbook(template_file)
    else:
        template_wb = openpyxl.load_workbook(template_file)

    template_ws = template_wb.active

    # Get existing records in template
    existing_records = set()
    for row in template_ws.iter_rows(
        min_row=2, max_row=template_ws.max_row, min_col=1, max_col=1
    ):
        if row[0].value:
            existing_records.add(str(row[0].value))

    # Process new records
    new_row = template_ws.max_row + 1
    for _, row in uploaded_profile_df.iterrows():
        if str(row[0]) not in existing_records and pd.notna(row[0]):
            # Add new record
            for col, value in enumerate(row, start=1):
                if pd.notna(value):
                    cell = template_ws.cell(row=new_row, column=col)
                    cell.value = value
                    # Copy formatting from the row above
                    source_cell = template_ws.cell(row=new_row - 1, column=col)
                    copy_cell_formatting(source_cell, cell)

            # Copy formulas from previous row
            copy_formulas(template_ws, new_row - 1, new_row)
            new_row += 1

    return template_wb


def process_audit_data(uploaded_audit_df, template_file):
    """Process audit data and update template"""
    # Handle both file object and file path
    if isinstance(template_file, str):
        template_wb = openpyxl.load_workbook(template_file)
    else:
        template_wb = openpyxl.load_workbook(template_file)

    template_ws = template_wb.active

    # Get headers from template
    template_headers = {}
    formula_columns = set()

    # Identify columns with formulas in the first data row
    first_data_row = 2
    for col in template_ws[first_data_row]:
        if col.value and isinstance(col.value, str) and col.value.startswith("="):
            formula_columns.add(col.column)

    # Get headers
    for col in template_ws[1]:
        template_headers[col.value] = col.column

    # Get existing records in template
    existing_records = set()
    for row in template_ws.iter_rows(
        min_row=2, max_row=template_ws.max_row, min_col=1, max_col=1
    ):
        if row[0].value:
            existing_records.add(str(row[0].value))

    # Process new records
    new_row = template_ws.max_row + 1
    for _, row in uploaded_audit_df.iterrows():
        if str(row[0]) not in existing_records and pd.notna(row[0]):
            # Map uploaded data to template columns
            for col_name, value in row.items():
                if pd.notna(value) and col_name in template_headers:
                    template_col = template_headers[col_name]
                    # Skip formula columns during data import
                    if template_col not in formula_columns:
                        cell = template_ws.cell(row=new_row, column=template_col)
                        cell.value = value
                        # Copy formatting from the row above
                        source_cell = template_ws.cell(
                            row=new_row - 1, column=template_col
                        )
                        copy_cell_formatting(source_cell, cell)

            # Copy formulas from previous row
            copy_formulas(template_ws, new_row - 1, new_row)
            new_row += 1

    return template_wb


def refresh_app():
    """Reset the app state"""
    st.session_state.processing_complete = False
    st.experimental_rerun()


def validate_file_path(file_path):
    """Validate if a file path exists and is accessible"""
    if not file_path:
        return False
    return os.path.exists(file_path) and os.path.isfile(file_path)


def main():
    st.title("Profile and Audit Data Update")

    if not st.session_state.processing_complete:
        # Add tabs for different input methods
        input_method = st.radio("Choose input method:", ["File Upload", "File Paths"])

        if input_method == "File Upload":
            # File uploads
            profile_file = st.file_uploader("Upload Profile Data (CSV)", type=["csv"])
            audit_file = st.file_uploader("Upload Audit Data (XLSX)", type=["xlsx"])
            profile_template = st.file_uploader(
                "Upload Profile Template (XLSX)", type=["xlsx"], key="profile_template"
            )
            audit_template = st.file_uploader(
                "Upload Audit Template (XLSX)", type=["xlsx"], key="audit_template"
            )

            files_ready = all(
                [profile_file, audit_file, profile_template, audit_template]
            )

        else:  # File Paths
            # File path inputs
            profile_path = st.text_input("Profile Data Path (CSV):")
            audit_path = st.text_input("Audit Data Path (XLSX):")
            profile_template_path = st.text_input("Profile Template Path (XLSX):")
            audit_template_path = st.text_input("Audit Template Path (XLSX):")

            # Validate paths
            paths_valid = all(
                validate_file_path(path)
                for path in [
                    profile_path,
                    audit_path,
                    profile_template_path,
                    audit_template_path,
                ]
                if path
            )

            if not paths_valid and any(
                [profile_path, audit_path, profile_template_path, audit_template_path]
            ):
                st.warning("One or more file paths are invalid or inaccessible.")

            files_ready = paths_valid and all(
                [profile_path, audit_path, profile_template_path, audit_template_path]
            )

        if st.button("Process Files") and files_ready:
            try:
                # Read files based on input method
                if input_method == "File Upload":
                    profile_df = pd.read_csv(profile_file)
                    audit_df = pd.read_excel(audit_file)
                    profile_template_data = profile_template
                    audit_template_data = audit_template
                else:
                    profile_df = pd.read_csv(profile_path)
                    audit_df = pd.read_excel(audit_path)
                    profile_template_data = profile_template_path
                    audit_template_data = audit_template_path

                # Store processed data in session state
                st.session_state.profile_buffer = io.BytesIO()
                st.session_state.audit_buffer = io.BytesIO()

                # Process Profile Data
                processed_profile = process_profile_data(
                    profile_df, profile_template_data
                )
                if processed_profile:
                    processed_profile.save(st.session_state.profile_buffer)
                    st.session_state.profile_buffer.seek(0)

                # Process Audit Data
                processed_audit = process_audit_data(audit_df, audit_template_data)
                if processed_audit:
                    processed_audit.save(st.session_state.audit_buffer)
                    st.session_state.audit_buffer.seek(0)

                st.session_state.processing_complete = True
                st.experimental_rerun()

            except Exception as e:
                st.error(f"An error occurred: {str(e)}")

    if st.session_state.processing_complete:
        # Generate filenames with current date
        current_date = datetime.now().strftime("%d%m%Y")
        profile_filename = f"ProfileData_{current_date}.xlsx"
        audit_filename = f"AuditData_{current_date}.xlsx"

        # Success messages and download buttons
        st.success("Profile data processed successfully!")
        st.download_button(
            label="Download Updated Profile Data",
            data=st.session_state.profile_buffer,
            file_name=profile_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.success("Audit data processed successfully!")
        st.download_button(
            label="Download Updated Audit Data",
            data=st.session_state.audit_buffer,
            file_name=audit_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Add refresh button at the bottom
        st.button("Start Fresh Update", on_click=refresh_app)


if __name__ == "__main__":
    main()
